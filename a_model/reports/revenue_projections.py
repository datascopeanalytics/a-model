import openpyxl

from .base import Report


class RevenueProjections(Report):
    report_name = 'revenue_projections.csv'
    gsheet_tab_name = 'Revenue Projections'

    def __init__(self, *args, **kwargs):
        super(RevenueProjections, self).__init__()
        self._cached_revenue_projections = None

    def get_revenue_projections(self):
        if self._cached_revenue_projections:
            return self._cached_revenue_projections
        self._cached_revenue_projections = revenue_projections = []
        worksheet = self.open_worksheet()
        max_col = openpyxl.cell.get_column_letter(worksheet.max_column)
        max_row = worksheet.max_row

        # aggregate all of the dates
        date_cells = self.iter_cells_in_row(1, 'C', max_col)
        dates = [self.get_date_from_cell(cell) for cell in date_cells]

        # iterate over all of the projected revenues for all clients
        cell_range = 'C2:%(max_col)s%(max_row)d' % locals()
        for row in worksheet.iter_rows(cell_range):
            for date, revenue_cell in zip(dates, row):
                revenue = self.get_float_from_cell(revenue_cell)
                if revenue > 0:
                    revenue_projections.append((date, revenue))
        return revenue_projections

    def __iter__(self):
        now = self.get_now()
        for date, amount in self.get_revenue_projections():
            if date < now:
                raise ValueError((
                    "Double check Revenue Projections spreadsheet. There is a "
                    "projected revenue in the past"
                ))
            yield date, amount
